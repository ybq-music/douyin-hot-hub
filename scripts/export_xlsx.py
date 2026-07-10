import json
import urllib.parse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "raw"
OUTPUT_DIR = ROOT / "outputs"

HEADERS = ["日期", "采集时间", "榜单", "分类", "排名", "名称", "作者/用户", "热度/指标", "链接", "原始数据"]


def latest_raw_dir():
    dirs = sorted([p for p in RAW_DIR.iterdir() if p.is_dir()])
    if not dirs:
        raise RuntimeError("raw 目录里还没有日期文件夹，请先运行 main.py")
    return dirs[-1]


def load_json(path):
    try:
        with path.open("r", encoding="utf-8") as f:
            text = f.read().strip()

        if not text:
            print(f"skip empty json: {path}")
            return {}

        return json.loads(text)

    except json.JSONDecodeError:
        print(f"skip invalid json: {path}")
        return {}


def find_list(value):
    if isinstance(value, list):
        return value
    if not isinstance(value, dict):
        return []

    for key in ["word_list", "rank_list", "music_list", "user_list", "room_list", "brand_list", "list", "items", "data"]:
        if key in value:
            found = find_list(value[key])
            if found:
                return found

    for child in value.values():
        found = find_list(child)
        if found:
            return found

    return []


def pick(item, paths):
    for path in paths:
        cur = item
        for part in path.split("."):
            if isinstance(cur, dict):
                cur = cur.get(part)
            else:
                cur = None
                break
        if cur not in (None, ""):
            return cur
    return ""


def as_text(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else str(value)


def make_url(chart, item):
    if chart == "抖音热榜":
        word = pick(item, ["word", "sentence", "title", "name"])
        return "https://www.douyin.com/search/" + urllib.parse.quote(as_text(word))

    if chart == "音乐榜":
        return pick(item, ["music_info.play_url.uri", "play_url.uri", "url", "share_url"])

    if chart == "明星榜":
        uid = pick(item, ["user_info.uid", "uid"])
        sec_uid = pick(item, ["user_info.sec_uid", "sec_uid"])
        if uid and sec_uid:
            return f"https://www.iesdouyin.com/share/user/{uid}?sec_uid={sec_uid}"

    if chart == "直播榜":
        room_id = pick(item, ["room.id", "room_id", "id"])
        if room_id:
            return f"https://webcast.amemv.com/webcast/reflow/{room_id}"

    return pick(item, ["url", "share_url", "schema"])


def normalize(chart, item, rank, date_text, captured_at, category=""):
    name = pick(item, [
        "word", "sentence", "title", "name",
        "music_info.title",
        "user_info.nickname",
        "user.nickname",
        "room.title",
    ])

    person = pick(item, [
        "music_info.author",
        "author",
        "user_info.nickname",
        "user.nickname",
        "nickname",
    ])

    metric = pick(item, [
        "hot_value", "hot", "score", "value",
        "rank_value", "view_count", "user_count",
    ])

    raw = json.dumps(item, ensure_ascii=False)
    if len(raw) > 30000:
        raw = raw[:30000]

    return [
        date_text,
        captured_at,
        chart,
        category,
        rank,
        as_text(name),
        as_text(person),
        as_text(metric),
        make_url(chart, item),
        raw,
    ]


def read_chart_rows(date_dir, chart, filename, category=""):
    path = date_dir / filename
    if not path.exists():
        return []

    data = load_json(path)
    items = find_list(data)
    captured_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    return [
        normalize(chart, item, i + 1, date_dir.name, captured_at, category)
        for i, item in enumerate(items)
        if isinstance(item, dict)
    ]


def write_sheet(wb, name, rows):
    ws = wb.create_sheet(name[:31])
    ws.append(HEADERS)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="366092")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"

    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18

    ws.column_dimensions["J"].width = 60


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    date_dir = latest_raw_dir()

    chart_rows = {
        "抖音热榜": read_chart_rows(date_dir, "抖音热榜", "hot-search.json"),
        "明星榜": read_chart_rows(date_dir, "明星榜", "hot-star.json"),
        "直播榜": read_chart_rows(date_dir, "直播榜", "hot-live.json"),
        "音乐榜": read_chart_rows(date_dir, "音乐榜", "hot-music.json"),
        "品牌榜": [],
    }

    brand_dir = date_dir / "brand"
    if brand_dir.exists():
        for path in sorted(brand_dir.glob("*.json")):
            chart_rows["品牌榜"].extend(
                read_chart_rows(date_dir, "品牌榜", f"brand/{path.name}", path.stem)
            )

    all_rows = []
    for rows in chart_rows.values():
        all_rows.extend(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    ws.append(["文件日期", date_dir.name])
    ws.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["说明", "本文件由 GitHub Actions 自动生成，记录当日所有抖音榜单。"])

    write_sheet(wb, "全部明细", all_rows)
    for name, rows in chart_rows.items():
        write_sheet(wb, name, rows)

    output = OUTPUT_DIR / f"douyin_ranks_{date_dir.name}.xlsx"
    wb.save(output)
    print(f"saved {output}")


if __name__ == "__main__":
    main()
